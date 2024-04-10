import openpyxl
from tqdm import tqdm
from openpyxl.utils.exceptions import InvalidFileException


def get_user_input(value):
    workbook = None

    try:
        # Open the Excel file
        workbook = openpyxl.load_workbook('input_file.xlsx', read_only=True)
        
        # Check if 'InputSheet' exists in the workbook
        if 'InputSheet' in workbook.sheetnames:
            sheet = workbook['InputSheet']

            # Get user input from cell B1
            user_input = sheet[f'A{value}'].value
            return user_input
        else:
            print("Error: Worksheet 'InputSheet' does not exist in the Excel file.")
            return None
    except InvalidFileException as e:
        print(f"Error: {e}")
        return None
    finally:
        # Close the workbook
        if workbook:
            workbook.close()


def perform_computation(user_input, value):
    square = user_input ** 2
    cubes =  user_input ** 3
    squareRoots =  user_input ** 1/2
    sum_of_all_three = square + cubes + squareRoots
   
    return [square, cubes, squareRoots,sum_of_all_three]

def write_output_to_excel(output, value):
    try:
        # Open the Excel file
        workbook = openpyxl.load_workbook('input_file.xlsx')
        sheet = workbook['InputSheet']
       
        sheet[f"B{value}"].value = output[0]
        sheet[f"C{value}"].value = output[1]   
        sheet[f"D{value}"].value = output[2]
        sheet[f"E{value}"].value = output[3]
        # Save the changes
        workbook.save('input_file.xlsx')
    except InvalidFileException as e:
        print(f"Error: {e}")
    finally:
        # Close the workbook
        if workbook:
            workbook.close()

if __name__ == "__main__":
    
    for value in tqdm(range(2,62)):
    # Get user input
        user_input = get_user_input(value)

        if user_input is not None:
            # Perform computation
            result = perform_computation(user_input,value)

            # Write output to Excel
            write_output_to_excel(result,value)

    print(f"All data records successfully.")
